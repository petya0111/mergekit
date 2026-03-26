import os
import re
import json
import datetime
import stat
from pathlib import Path
from difflib import SequenceMatcher
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

import pandas as pd

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def set_file_readonly(filepath: Path, readonly: bool = True):
    """Set or remove read-only flag on a file."""
    try:
        if readonly:
            # Remove write permissions
            os.chmod(filepath, stat.S_IRUSR | stat.S_IRGRP | stat.S_IROTH)
        else:
            # Add write permissions
            os.chmod(filepath, stat.S_IRUSR | stat.S_IWUSR | stat.S_IRGRP | stat.S_IROTH)
    except Exception:
        pass


def is_file_readonly(filepath: Path) -> bool:
    """Check if a file is read-only (no write permission)."""
    try:
        if not filepath.exists():
            return False
        mode = os.stat(filepath).st_mode
        return not (mode & stat.S_IWUSR)
    except Exception:
        return False


# -------------------------
# Persistent storage paths
# -------------------------
BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
ORDERS_DIR = None  # Не се използва - потребителят избира къде да записва
PROTOCOLS_DIR = None  # Ще се зададе от потребителя
SETTINGS_FILE = DATA_DIR / "settings.json"


# -------------------------
# Helpers: header detection
# -------------------------
def normalize(s: str) -> str:
    """Normalize a header or key: lowercase, remove punctuation and collapse whitespace.
    This makes matching more robust for variants like 'шир./вис.' vs 'шир/вис'.
    """
    s = str(s).strip().lower()
    # replace any non-word/digit characters with a single space (keeps Cyrillic/latin letters and digits)
    s = re.sub(r"[^\w\d]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s
def canonical_code(x) -> str:
    """
    Канонизира код/артикул за сравнение между таблици.
    - маха .0 (ако Excel го е прочел като float)
    - маха интервали / NBSP
    - прави upper()
    """
    if pd.isna(x):
        return ""
    # ако е float и е цяло число -> int
    if isinstance(x, float) and x.is_integer():
        x = int(x)
    s = str(x).strip()
    # махни .0 ако идва от float в текст
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    # махни интервали и NBSP
    s = re.sub(r"[\s\u00A0\u202F]+", "", s)
    return s.upper()


def fuzzy_match_best(item: str, candidates: list, threshold: float = 0.6):
    """
    Find the best fuzzy match for `item` among `candidates`.
    Returns (best_candidate, score) or (None, 0) if no match above threshold.
    """
    item_lower = item.lower()
    best = None
    best_score = 0
    for c in candidates:
        if not isinstance(c, str):
            continue
        score = SequenceMatcher(None, item_lower, c.lower()).ratio()
        if score > best_score:
            best_score = score
            best = c
    if best_score >= threshold:
        return best, best_score
    return None, 0


def to_float(x):
    """
    Подобрен float parser:
    - приема '0,85', '0.85', '0,85 лв', '€0.85', '1 234,56'
    - вади първото число от текста
    """
    if pd.isna(x):
        return None
    if isinstance(x, (int, float)):
        return float(x)

    s = str(x).strip()
    if not s:
        return None

    # махни интервали/хилядарни разделители
    s = re.sub(r"[\s\u00A0\u202F]+", "", s)

    # вземи първото число (с , или .)
    m = re.search(r"[-+]?\d+(?:[.,]\d+)?", s)
    if not m:
        return None

    num = m.group(0).replace(",", ".")
    try:
        return float(num)
    except Exception:
        return None


def parse_qty_range_from_header(header: str):
    """
    Извлича (min,max) от заглавия като:
      '1 000 - 1 999', '2000-2999', '1 000 –1 999 бр'
    """
    if header is None:
        return None
    s = str(header).replace("–", "-").replace("—", "-")

    parts = re.findall(r"\d[\d\s\u00A0\u202F\.,]*\d|\d+", s)
    nums = []
    for p in parts:
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            try:
                nums.append(int(digits))
            except Exception:
                pass

    if len(nums) >= 2:
        return nums[0], nums[1]
    return None


def detect_range_columns(df_prices: pd.DataFrame):
    ranges = []
    for c in df_prices.columns:
        rng = parse_qty_range_from_header(c)
        if rng:
            ranges.append((rng[0], rng[1], c))
    ranges.sort(key=lambda x: (x[0], x[1]))
    return ranges


def resolve_unit_price_from_ranges(qty: int, price_row: pd.Series, ranges):
    if not ranges:
        return None

    for min_q, max_q, col in ranges:
        if min_q <= qty <= max_q:
            return to_float(price_row.get(col))

    # qty над последния диапазон -> ползвай последния
    last_min, last_max, last_col = ranges[-1]
    if qty > last_max:
        return to_float(price_row.get(last_col))

    return None


def find_column(df: pd.DataFrame, candidates):
    """
    Find a column by checking if any candidate substring appears in the header.
    Returns column name or raises ValueError.
    """
    cols = list(df.columns)
    norm_cols = {c: normalize(c) for c in cols}
    cand_norm = [normalize(x) for x in candidates]

    for c in cols:
        h = norm_cols[c]
        if any(cn in h for cn in cand_norm):
            return c
    raise ValueError(f"Не намерих колона за: {candidates}. Налични колони: {cols}")


def to_int(x):
    if pd.isna(x):
        return None
    if isinstance(x, (int,)):
        return int(x)
    try:
        s = str(x).strip()
        # remove thousands separators (spaces, NBSP)
        s = re.sub(r"[\s\u00A0\u202F]+", "", s)
        s = s.replace(",", ".")
        v = float(s)
        return int(round(v))
    except Exception:
        return None




def excel_cell_to_string(x) -> str:
    """Return value exactly as string from Excel (no parsing/formatting)."""
    if pd.isna(x):
        return ""
    return str(x).strip()


# -------------------------
# NEW: price ranges like "1 000 - 1 999"
# -------------------------
def parse_qty_range_from_header(header):
    """
    Extract (min,max) from headers like:
      "1 000 -1 999", "2 000 - 2 999", "1000-1999", "1 000 – 1 999"
      Also handles single numbers like 50000 -> (50000, 59999)
    Returns (min,max) or None.
    """
    if header is None:
        return None
    
    # If header is already a number
    if isinstance(header, (int, float)) and not pd.isna(header):
        qty = int(header)
        # Single number: treat as exact quantity with some range
        # e.g., 50000 means 50000-59999, 60000 means 60000-79999, etc.
        return (qty, qty)  # Will be handled specially in resolve_unit_price
    
    s = str(header).replace("–", "-").replace("—", "-")

    # find numbers possibly containing spaces/dots/commas (thousand separators)
    parts = re.findall(r"\d[\d\s\u00A0\u202F\.,]*\d|\d+", s)
    nums = []
    for p in parts:
        digits = re.sub(r"[^0-9]", "", p)
        if digits:
            try:
                nums.append(int(digits))
            except Exception:
                pass
    if len(nums) >= 2:
        return nums[0], nums[1]
    elif len(nums) == 1:
        # Single number in header text
        return (nums[0], nums[0])
    return None


def detect_range_columns(df_prices: pd.DataFrame):
    """
    Finds range-price columns in prices sheet and returns:
      [(min_qty, max_qty, colname), ...] sorted by min_qty.
    Also detects single-number columns like 50000, 60000.
    """
    ranges = []
    for c in df_prices.columns:
        rng = parse_qty_range_from_header(c)
        if rng:
            ranges.append((rng[0], rng[1], c))
    ranges.sort(key=lambda x: (x[0], x[1]))
    return ranges


def resolve_unit_price_from_ranges(qty: int, price_row: pd.Series, ranges):
    """
    Pick unit price by finding the column whose (min<=qty<=max).
    For single-number columns, picks the closest one that doesn't exceed qty.
    If qty is above the last range, use the last range's price (fallback).
    If qty is below the first range, use the first range's price (fallback).
    """
    if not ranges:
        return None

    # First, try exact range match
    for min_q, max_q, col in ranges:
        if min_q <= qty <= max_q:
            price = to_float(price_row.get(col))
            if price is not None:
                return price

    # For single-number columns (min==max), find the best match
    # Pick the highest single-number column that is <= qty
    single_cols = [(min_q, col) for min_q, max_q, col in ranges if min_q == max_q]
    if single_cols:
        # Sort by quantity descending
        single_cols.sort(key=lambda x: x[0], reverse=True)
        for threshold, col in single_cols:
            if qty >= threshold:
                price = to_float(price_row.get(col))
                if price is not None:
                    return price

    # fallback: qty above last range -> use last available price
    last_min, last_max, last_col = ranges[-1]
    if qty > last_max:
        price = to_float(price_row.get(last_col))
        if price is not None:
            return price

    # qty below first range -> use first available price
    first_min, first_max, first_col = ranges[0]
    if qty < first_min:
        price = to_float(price_row.get(first_col))
        if price is not None:
            return price
    
    # Try to find ANY non-null price in the ranges
    for min_q, max_q, col in ranges:
        price = to_float(price_row.get(col))
        if price is not None:
            return price

    return None


# -------------------------
# Excel reading
# -------------------------
def read_excel_any(path: str) -> pd.DataFrame:
    """
    Read .xls/.xlsx with pandas.
    .xls requires xlrd==2.0.1 installed.
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".xls":
            return pd.read_excel(path, engine="xlrd")
        else:
            return pd.read_excel(path, engine="openpyxl")
    except ImportError as e:
        raise ImportError(
            "Липсва библиотека за четене. Инсталирай:\n"
            "pip install pandas openpyxl xlrd==2.0.1\n\n"
            f"Оригинална грешка: {e}"
        )
    except Exception as e:
        raise RuntimeError(f"Не успях да прочета файла: {path}\nГрешка: {e}")


def merge_order_and_prices(order_path: str, prices_path_or_df) -> pd.DataFrame:
    """Merge order with prices.
    
    Args:
        order_path: Path to the order Excel file
        prices_path_or_df: Either a path to a prices Excel file, or a DataFrame with prices
    """
    df_order = read_excel_any(order_path)
    
    # Accept either path or DataFrame for prices
    if isinstance(prices_path_or_df, pd.DataFrame):
        df_prices = prices_path_or_df
    else:
        df_prices = read_excel_any(prices_path_or_df)

    DEBUG = os.environ.get("MERGE_DEBUG", "0") in ("1", "true", "True")
    if DEBUG:
        print("[DEBUG] Order columns:", list(df_order.columns))
        print("[DEBUG] Prices columns:", list(df_prices.columns))

    # Find order columns
    col_order_no = find_column(df_order, ["номер на поръчка", "поръчка", "Purchase Order"])
    col_item = find_column(df_order, ["име на артикул", "артикул", "продукт", "Item Number"])
    col_qty = find_column(df_order, ["заявени бройки", "бройки", "количество", "Quantity Ordered"])
    col_date = find_column(df_order, ["дата на доставка", "доставка", "delivery", "Due Date"])
    
    # Try to find Purchase Order Line column for sorting
    try:
        col_order_line = find_column(df_order, ["Purchase Order Line", "ред на поръчка", "order line"])
    except Exception:
        col_order_line = None

    # Find prices columns
    p_item = find_column(df_prices, ["код АЛ филтър", "артикул", "item"])

    try:
        p_tl = find_column(df_prices, ["технологичен лист", "ТЛ", "tech"])
    except Exception:
        p_tl = None

    try:
        p_size = find_column(
            df_prices,
            ["размер", "шир./вис.", "шир/вис", "ширина/височина", "ширина/вис", "size", "width/height"]
        )
    except Exception:
        p_size = None
        # fallback: try to find a header containing both шир and вис
        for c in df_prices.columns:
            try:
                h = normalize(c)
                if "шир" in h and "вис" in h:
                    p_size = c
                    break
            except Exception:
                continue

    try:
        p_mat = find_column(df_prices, ["материал", "material"])
    except Exception:
        p_mat = None

    # Detect range columns like "1 000 - 1 999", "2 000 - 2 999", ...
    range_cols = detect_range_columns(df_prices)
    if not range_cols:
        raise ValueError(
            "Не открих ценови колони тип диапазон (напр. '1 000 - 1 999'). "
            "Провери заглавията в таблица 'Цени'."
        )

    # Build lookup: keep full row so we can read price from range columns
    # Use multiple keys for better matching: original, normalized, and canonical
    # When there are multiple rows for the same article, keep the one with more filled prices
    prices_lookup = {}
    
    def count_filled_prices(row):
        """Count how many price range columns have valid (non-NaN) values."""
        count = 0
        for _, _, col in range_cols:
            val = row.get(col)
            if not pd.isna(val):
                count += 1
        return count
    
    for _, pr in df_prices.iterrows():
        name = pr.get(p_item)
        if pd.isna(name):
            continue
        name = str(name).strip()
        name_norm = normalize(name)
        name_canon = canonical_code(name)
        info = {
            "row": pr,
            "Технологичен лист": "" if p_tl is None or pd.isna(pr.get(p_tl)) else str(pr.get(p_tl)).strip(),
            "Размер": "" if p_size is None or pd.isna(pr.get(p_size)) else str(pr.get(p_size)).strip(),
            "Материал": "" if p_mat is None or pd.isna(pr.get(p_mat)) else str(pr.get(p_mat)).strip(),
        }
        
        # Only update if this is a new key or if this row has more or equal filled price columns
        # (equal means later file wins - useful when combining multiple price files)
        new_count = count_filled_prices(pr)
        for key in [name, name_norm, name_canon]:
            if key is None:
                continue
            existing = prices_lookup.get(key)
            if existing is None:
                prices_lookup[key] = info
            else:
                # Compare: keep the row with more filled prices, or the later one if equal
                old_count = count_filled_prices(existing["row"])
                if new_count >= old_count:
                    prices_lookup[key] = info

    if DEBUG:
        print(f"[DEBUG] Built prices_lookup with {len(prices_lookup)} keys. Sample keys: {list(prices_lookup.keys())[:6]}")

    # Merge
    out_rows = []
    line_no = 0

    for _, r in df_order.iterrows():
        order_no = r.get(col_order_no)
        item = r.get(col_item)
        qty = r.get(col_qty)
        ddate = r.get(col_date)
        
        # Get the actual order line number if available
        order_line = None
        if col_order_line is not None:
            order_line = r.get(col_order_line)
            if pd.notna(order_line):
                order_line = int(order_line)

        if pd.isna(order_no) or pd.isna(item):
            continue

        order_no = str(order_no).strip()
        item = str(item).strip()
        item_norm = normalize(item)
        item_canon = canonical_code(item)

        qty_i = to_int(qty)
        if qty_i is None:
            continue

        line_no += 1
        # Use actual order line number if available, otherwise use counter
        actual_line = order_line if order_line is not None else line_no
        order_ref = f"{order_no}-{actual_line}"

        # find price row - STRICT matching by canonical code only
        # (артикулният номер от поръчката трябва да съвпада точно с "код АЛ филтър")
        price_info = None
        exact_match = False
        
        # Try exact canonical code match first
        if item_canon and item_canon in prices_lookup:
            price_info = prices_lookup.get(item_canon)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact canonical match: '{item}' (canon='{item_canon}')")
        
        # Try exact normalized match
        if not exact_match and item_norm in prices_lookup:
            price_info = prices_lookup.get(item_norm)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact normalized match: '{item}' (norm='{item_norm}')")
        
        # Try exact original match
        if not exact_match and item in prices_lookup:
            price_info = prices_lookup.get(item)
            exact_match = True
            if DEBUG:
                print(f"[DEBUG] Exact original match: '{item}'")
        
        # Try partial code match: if order item contains the price code or vice versa
        if not exact_match:
            for price_key, price_val in prices_lookup.items():
                if not isinstance(price_key, str) or not price_key:
                    continue
                price_canon = canonical_code(price_key)
                # Check if item code contains price code or price code contains item code
                if price_canon and item_canon:
                    if price_canon in item_canon or item_canon in price_canon:
                        price_info = price_val
                        exact_match = True
                        if DEBUG:
                            print(f"[DEBUG] Partial code match: '{item}' (canon='{item_canon}') contains/in '{price_key}' (canon='{price_canon}')")
                        break
        
        if not exact_match and DEBUG:
            print(f"[DEBUG] No exact match for '{item}' (canon='{item_canon}') - Ед. Цена and Сума will be empty")

        unit_price = None
        size = ""
        tl = ""
        mat = ""

        if price_info and exact_match:
            unit_price = resolve_unit_price_from_ranges(qty_i, price_info["row"], range_cols)
            size = price_info.get("Размер", "") or ""
            tl = price_info.get("Технологичен лист", "") or ""
            mat = price_info.get("Материал", "") or ""
            if DEBUG and unit_price is None:
                # show raw values in each range column for troubleshooting
                raw_vals = [(c, price_info["row"].get(c)) for _, _, c in range_cols]
                print(f"[DEBUG] Item '{item}' matched but unit_price=None. qty={qty_i}. Raw range values: {raw_vals[:6]}...")
        else:
            # No exact match - leave price fields empty
            if DEBUG:
                print(f"[DEBUG] Item '{item}' - no exact match, Ед. Цена and Сума will be empty.")

        total = round(unit_price * qty_i, 2) if unit_price is not None else None

        out_rows.append({
            "Артикул": item,
            "Размер": size,
            "Бройки": qty_i,
            "Ед. Цена": "" if unit_price is None else unit_price,
            "Сума": "" if total is None else total,
            "Номер на поръчка и ред": order_ref,
            "Дата на доставка": excel_cell_to_string(ddate),  # keep as string
            "Технологичен лист": tl,
            "Материал": mat,
        })

    result_df = pd.DataFrame(out_rows)
    
    # Sort by order line number (extracted from "Номер на поръчка и ред")
    if not result_df.empty and "Номер на поръчка и ред" in result_df.columns:
        def extract_line_number(ref):
            """Extract the line number from 'B26801-5' format."""
            try:
                parts = str(ref).split("-")
                if len(parts) >= 2:
                    return int(parts[-1])
            except:
                pass
            return 0
        
        result_df["_sort_key"] = result_df["Номер на поръчка и ред"].apply(extract_line_number)
        result_df = result_df.sort_values("_sort_key").drop(columns=["_sort_key"]).reset_index(drop=True)
    
    return result_df


def _apply_date_format_xlsx(path: Path, header_name: str = "Дата на доставка"):
    """
    Post-process an .xlsx file to ensure the column with header_name contains real datetimes
    and has a sensible Excel number format.
    NOTE: You wanted date as string, so this is typically NOT needed for your generated orders.
    It's kept only because your protocol logic uses it.
    """
    if load_workbook is None:
        return
    try:
        wb = load_workbook(filename=str(path))
        ws = wb.active
        header_col = None
        for cell in ws[1]:
            try:
                if str(cell.value).strip() == header_name:
                    header_col = cell.column
                    break
            except Exception:
                continue
        if header_col is None:
            wb.close()
            return

        for row in ws.iter_rows(min_row=2, min_col=header_col, max_col=header_col):
            cell = row[0]
            val = cell.value
            if val is None:
                continue
            import datetime as _dt
            if isinstance(val, (_dt.datetime, _dt.date)):
                cell.number_format = 'yyyy-mm-dd'
                continue

            try:
                parsed = pd.to_datetime(val, dayfirst=True, errors='coerce')
            except Exception:
                parsed = pd.to_datetime(val, errors='coerce')
            if pd.isna(parsed):
                continue
            try:
                cell.value = parsed.to_pydatetime()
            except Exception:
                continue
            cell.number_format = 'yyyy-mm-dd'

        wb.save(filename=str(path))
        wb.close()
    except Exception:
        try:
            wb.close()
        except Exception:
            pass


def _apply_inventory_colors_xlsx(path: Path):
    """Оцветява редовете в записания Excel спрямо 'Статус наличност'."""
    if load_workbook is None:
        return
    from openpyxl.styles import PatternFill

    STATUS_FILLS = {
        "Достатъчно":   PatternFill("solid", fgColor="D4EDDA"),
        "Няма":         PatternFill("solid", fgColor="F8D7DA"),
        "nostock":      PatternFill("solid", fgColor="CCE5FF"),   # Не в склада
        "difftl":       PatternFill("solid", fgColor="E8D5F5"),   # Различен ТЛ
        "low":          PatternFill("solid", fgColor="FFF3CD"),   # Недостатъчно
    }

    def _fill_for_status(s: str):
        if not s:
            return None
        if s == "Достатъчно":
            return STATUS_FILLS["Достатъчно"]
        if s == "Няма":
            return STATUS_FILLS["Няма"]
        if s.startswith("Недостатъчно"):
            return STATUS_FILLS["low"]
        if s.startswith("Различен ТЛ"):
            return STATUS_FILLS["difftl"]
        if s == "Не в склада":
            return STATUS_FILLS["nostock"]
        return None

    try:
        wb = load_workbook(filename=str(path))
        ws = wb.active
        # Find header positions
        status_col = None
        max_col = ws.max_column
        for cell in ws[1]:
            try:
                if str(cell.value).strip() == "Статус наличност":
                    status_col = cell.column
                    break
            except Exception:
                continue
        if status_col is None:
            wb.close()
            return

        for row in ws.iter_rows(min_row=2):
            status_val = str(row[status_col - 1].value or "").strip()
            fill = _fill_for_status(status_val)
            if fill:
                for cell in row:
                    cell.fill = fill

        wb.save(filename=str(path))
        wb.close()
    except Exception:
        try:
            wb.close()
        except Exception:
            pass


def load_settings():
    try:
        if SETTINGS_FILE.exists():
            with open(SETTINGS_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}


def save_settings(s: dict):
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(SETTINGS_FILE, "w", encoding="utf-8") as f:
            json.dump(s, f, ensure_ascii=False, indent=2)
    except Exception:
        pass


def set_protocols_dir(path_str: str):
    """Set protocols directory (no persistence - user must choose each session)."""
    global PROTOCOLS_DIR
    try:
        p = Path(path_str).expanduser().resolve()
    except Exception:
        p = Path(path_str)
    p.mkdir(parents=True, exist_ok=True)
    PROTOCOLS_DIR = p


def ensure_dirs():
    """Създава директориите за протоколи ако са избрани."""
    if PROTOCOLS_DIR is not None:
        try:
            PROTOCOLS_DIR.mkdir(parents=True, exist_ok=True)
        except Exception:
            pass


def week_key_from_date(d):
    if pd.isna(d) or d == "":
        return "protocol_undated"
    if isinstance(d, str):
        s = d.strip()
        # Try ISO format first (YYYY-MM-DD) - must not use dayfirst for this
        if re.match(r"^\d{4}-\d{2}-\d{2}", s):
            dt = pd.to_datetime(s, errors="coerce")
        else:
            # For other formats like DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY use dayfirst
            try:
                dt = pd.to_datetime(s, dayfirst=True, errors="coerce")
            except Exception:
                dt = pd.to_datetime(s, errors="coerce")
    else:
        dt = pd.to_datetime(d, errors="coerce")

    if pd.isna(dt):
        return "protocol_undated"
    iso = dt.isocalendar()
    return f"protocol_{iso.year}_w{iso.week}"


def append_to_protocol(protocol_key: str, df_rows: pd.DataFrame, source_filename: str):
    ensure_dirs()
    prot_xlsx = PROTOCOLS_DIR / f"{protocol_key}.xlsx"
    
    # Check if protocol is closed (read-only or _CLOSED in name)
    if "_CLOSED" in protocol_key or is_file_readonly(prot_xlsx):
        raise RuntimeError(f"Протокол {protocol_key} е приключен. Нови редове не могат да се добавят.")

    cols = ["Артикул", "Размер", "Бройки", "Ед. Цена", "Сума", "Номер на поръчка и ред", "Дата на доставка", "Технологичен лист", "Материал"]
    out = df_rows.copy()
    for c in cols:
        if c not in out.columns:
            out[c] = ""
    out = out[cols]

    if prot_xlsx.exists():
        try:
            existing = pd.read_excel(prot_xlsx, engine="openpyxl")
            
            # Remove duplicates: if "Номер на поръчка и ред" already exists, replace with new data
            if "Номер на поръчка и ред" in existing.columns and "Номер на поръчка и ред" in out.columns:
                # Get the order refs from new data
                new_refs = set(out["Номер на поръчка и ред"].dropna().astype(str).tolist())
                # Keep only rows from existing that are NOT in new data
                existing_filtered = existing[~existing["Номер на поръчка и ред"].astype(str).isin(new_refs)]
                new_all = pd.concat([existing_filtered, out], ignore_index=True)
            else:
                new_all = pd.concat([existing, out], ignore_index=True)
        except Exception:
            new_all = out
    else:
        new_all = out

    # If you really want protocols to have real date cells, keep this.
    if "Дата на доставка" in new_all.columns:
        try:
            new_all["Дата на доставка"] = pd.to_datetime(new_all["Дата на доставка"], errors="coerce")
        except Exception:
            pass

    new_all.to_excel(prot_xlsx, index=False)
    try:
        _apply_date_format_xlsx(prot_xlsx, header_name="Дата на доставка")
    except Exception:
        pass


# -------------------------
# Inventory (Наличности) Functions
# -------------------------

def load_inventory(inventory_path: str) -> pd.DataFrame:
    """Load inventory file with available labels."""
    df = read_excel_any(inventory_path)
    return df


def _norm_tl(val) -> str:
    """
    Normalise a tech-list value to a clean string for matching.
    Handles float/int (22872.0 → "22872"), BOM, whitespace.
    Returns "" for NaN/None/empty.
    """
    if val is None or (isinstance(val, float) and pd.isna(val)):
        return ""
    s = str(val).lstrip('\ufeff').strip()
    # "22872.0" → "22872"
    if s.endswith('.0') and s[:-2].lstrip('-').isdigit():
        s = s[:-2]
    return s


def _find_inv_col(columns, candidates: list) -> str:
    """
    Robustly find a column name from a list of candidates.
    Strips BOM, extra whitespace, and compares case-insensitively.
    Also does partial/contains matching as fallback.
    """
    # Normalize helper: strip BOM + whitespace, lower, collapse spaces
    def norm(s):
        return str(s).lstrip('\ufeff').strip().lower().replace('\xa0', ' ')

    norm_candidates = [norm(c) for c in candidates]

    # Pass 1 — exact match after normalising
    for col in columns:
        if norm(col) in norm_candidates:
            return col

    # Pass 2 — partial/contains match (e.g. "тех.лист" inside "тех.лист (номер)")
    for col in columns:
        n = norm(col)
        for cand in norm_candidates:
            if cand in n or n in cand:
                return col

    return None


def check_inventory(df_order: pd.DataFrame, df_inventory: pd.DataFrame) -> pd.DataFrame:
    """
    Check order against inventory and add columns:
    - 'Налични' - how many are available in inventory
    - 'Статус наличност':
        'Достатъчно'        - намерен по артикул/ТЛ, налични >= поръчани
        'Недостатъчно (N)'  - намерен по артикул/ТЛ, 0 < налични < поръчани
        'Няма'              - намерен по артикул/ТЛ, налични == 0
        'Различен ТЛ (X)'  - артикулът е в inventory, но ТЛ не съвпада
        'Не в склада'       - артикулът не фигурира в inventory
        ''                  - няма нито артикул, нито ТЛ в поръчката

    Matching priority:
      1. Артикул + ТЛ съвпадат              → пълно съвпадение
      2. Само ТЛ съвпада                    → пълно съвпадение
      3. Само артикул съвпада, ТЛ различен  → 'Различен ТЛ'
      4. Нищо не съвпада                    → 'Не в склада'
    """
    # Find columns in inventory
    art_col = _find_inv_col(df_inventory.columns,
                            ['ime', 'име', 'артикул', 'name', 'item', 'item number', 'код'])
    tl_col  = _find_inv_col(df_inventory.columns,
                            ['тех.лист', 'технологичен лист', 'тл', 'tech list', 'tl'])
    qty_col = _find_inv_col(df_inventory.columns,
                            ['налични бройки', 'налични', 'количество', 'qty', 'quantity', 'бройки'])

    if qty_col is None:
        df_order['Налични'] = ''
        df_order['Статус наличност'] = 'Няма данни'
        return df_order

    # art_lookup: { UPPER(article) -> {'qty': int, 'tl': str} }
    # tl_lookup:  { norm(tl)       -> int qty }
    art_lookup: dict = {}
    tl_lookup:  dict = {}

    for _, row in df_inventory.iterrows():
        qty = row.get(qty_col)
        try:
            qty = int(float(qty)) if not pd.isna(qty) else 0
        except Exception:
            qty = 0

        inv_tl = _norm_tl(row.get(tl_col)) if tl_col else ''

        if art_col:
            art_str = _norm_tl(row.get(art_col))
            if art_str:
                art_lookup[art_str.upper()] = {'qty': qty, 'tl': inv_tl}

        if inv_tl:
            tl_lookup[inv_tl] = qty

    # Find columns in order df
    order_art_col = _find_inv_col(df_order.columns,
                                  ['артикул', 'item number', 'item', 'код', 'product'])
    order_tl_col  = _find_inv_col(df_order.columns,
                                  ['технологичен лист', 'тл', 'тех лист', 'tech list', 'tl'])

    available_list = []
    status_list    = []

    for _, row in df_order.iterrows():
        art_raw     = row.get(order_art_col, '') if order_art_col else ''
        tl_raw      = row.get(order_tl_col,  '') if order_tl_col  else ''
        ordered_qty = row.get('Бройки', 0)

        art_str = _norm_tl(art_raw)
        tl_str  = _norm_tl(tl_raw)

        # --- Priority 1 & 2: match by TL (exact) ---
        if tl_str and tl_str in tl_lookup:
            available = tl_lookup[tl_str]
            # Full match by TL
            if available == 0:
                available_list.append(0)
                status_list.append('Няма')
            elif available >= ordered_qty:
                available_list.append(available)
                status_list.append('Достатъчно')
            else:
                available_list.append(available)
                status_list.append(f'Недостатъчно ({available})')
            continue

        # --- Priority 2b: match by article ---
        if art_str and art_str.upper() in art_lookup:
            entry     = art_lookup[art_str.upper()]
            available = entry['qty']
            inv_tl    = entry['tl']

            # Check if TL matches (or order has no TL → accept)
            tl_matches = (not tl_str) or (tl_str == inv_tl)

            if not tl_matches:
                # Артикулът е намерен но ТЛ е различен
                available_list.append(available)
                status_list.append(f'Различен ТЛ ({inv_tl})')
                continue

            # TL matches or order has no TL → full match
            if available == 0:
                available_list.append(0)
                status_list.append('Няма')
            elif available >= ordered_qty:
                available_list.append(available)
                status_list.append('Достатъчно')
            else:
                available_list.append(available)
                status_list.append(f'Недостатъчно ({available})')
            continue

        # --- Not found at all ---
        if art_str or tl_str:
            available_list.append('')
            status_list.append('Не в склада')
        else:
            available_list.append('')
            status_list.append('')

    df_order['Налични'] = available_list
    df_order['Статус наличност'] = status_list

    return df_order

    return df_order

def reserve_inventory(df_order: pd.DataFrame, inventory_path: str, order_ref: str = None) -> tuple:
    """
    Reserve items from inventory based on order.
    Updates the inventory file by:
    - Subtracting ordered quantities from available (preserves cell colours/formatting via openpyxl)
    - Adding reservation note to history column

    Returns: (success_count, failed_list)
    """
    from datetime import datetime
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    df_inventory = read_excel_any(inventory_path)

    # Robust column finding (handles BOM / Windows extra spaces)
    art_col     = _find_inv_col(df_inventory.columns,
                                ['име', 'артикул', 'name', 'item', 'item number', 'код'])
    tl_col      = _find_inv_col(df_inventory.columns,
                                ['тех.лист', 'технологичен лист', 'тл', 'tech list', 'tl'])
    qty_col     = _find_inv_col(df_inventory.columns,
                                ['налични бройки', 'налични', 'количество', 'qty', 'quantity', 'бройки'])
    history_col = _find_inv_col(df_inventory.columns,
                                ['поръчка, бройка, дата', 'поръчка', 'история', 'history'])

    if qty_col is None:
        return (0, ['Липсват необходими колони в файла с наличности'])

    col_list        = list(df_inventory.columns)
    qty_col_idx     = col_list.index(qty_col) + 1
    history_col_idx = col_list.index(history_col) + 1 if history_col else None

    # Find columns in the order df
    order_art_col = _find_inv_col(df_order.columns,
                                  ['артикул', 'item number', 'item', 'код', 'product'])
    order_tl_col  = _find_inv_col(df_order.columns,
                                  ['технологичен лист', 'тл', 'тех лист', 'tech list', 'tl'])

    success_count = 0
    failed_list   = []

    # Build TWO maps: by article name and by tech list (shared entry objects)
    art_map: dict = {}   # { UPPER(article) -> entry }
    tl_map:  dict = {}   # { norm(tl)       -> entry }

    for i, row in df_inventory.iterrows():
        qty = row.get(qty_col)
        try:
            qty = int(float(qty)) if not pd.isna(qty) else 0
        except Exception:
            qty = 0
        entry = {'df_idx': i, 'qty': qty}

        if art_col:
            art_str = _norm_tl(row.get(art_col))
            if art_str:
                art_map[art_str.upper()] = entry
        if tl_col:
            tl_str = _norm_tl(row.get(tl_col))
            if tl_str:
                tl_map[tl_str] = entry

    def _find_entry(order_row):
        art_raw = order_row.get(order_art_col, '') if order_art_col else ''
        tl_raw  = order_row.get(order_tl_col,  '') if order_tl_col  else ''
        art_str = _norm_tl(art_raw)
        tl_str  = _norm_tl(tl_raw)
        if art_str:
            e = art_map.get(art_str.upper())
            if e is not None:
                return e, art_str
        if tl_str:
            e = tl_map.get(tl_str)
            if e is not None:
                return e, tl_str
        return None, (art_str or tl_str or '')

    # Collect updates: { df_idx: (new_qty, new_history) }
    updates: dict = {}

    for _, order_row in df_order.iterrows():
        ordered_qty = order_row.get('Бройки', 0)
        order_no    = order_row.get('Номер на поръчка и ред', order_ref or '')

        # Дата на доставка от поръчката (вместо днешна дата)
        delivery_date_raw = order_row.get('Дата на доставка', '')
        try:
            if delivery_date_raw and not pd.isna(delivery_date_raw):
                dt = pd.to_datetime(delivery_date_raw, dayfirst=True, errors='coerce')
                delivery_date = dt.strftime('%d.%m') if not pd.isna(dt) else ''
            else:
                delivery_date = ''
        except Exception:
            delivery_date = str(delivery_date_raw).strip()

        entry, label = _find_entry(order_row)

        if not label or ordered_qty == 0:
            continue

        if entry is None:
            failed_list.append(f'{label}: Не е намерен в наличности')
            continue

        df_idx      = entry['df_idx']
        current_qty = entry['qty']

        if current_qty < ordered_qty:
            failed_list.append(f'{label}: Недостатъчно ({current_qty} < {ordered_qty})')
            reserved = current_qty
            new_qty  = 0
        else:
            reserved = ordered_qty
            new_qty  = current_qty - ordered_qty

        # Update in-memory qty so repeated rows accumulate correctly
        entry['qty'] = new_qty

        new_entry = f'{order_no} {reserved} {delivery_date}'.strip()
        if df_idx in updates:
            prev_qty, prev_hist = updates[df_idx]
            updates[df_idx] = (new_qty, f'{prev_hist}; {new_entry}')
        else:
            current_history = df_inventory.loc[df_idx, history_col] if history_col else ''
            if pd.isna(current_history) or str(current_history).strip() == '':
                hist = new_entry
            else:
                hist = f'{current_history}; {new_entry}'
            updates[df_idx] = (new_qty, hist)

        success_count += 1

    # ----------------------------------------------------------------
    # Write back via openpyxl to PRESERVE cell colours / formatting
    # ----------------------------------------------------------------
    wb = load_workbook(inventory_path)
    ws = wb.active

    for df_idx, (new_qty, new_hist) in updates.items():
        wb_row = df_idx + 2  # +1 for header, +1 for 1-based indexing
        ws.cell(row=wb_row, column=qty_col_idx).value = new_qty
        if history_col_idx is not None:
            ws.cell(row=wb_row, column=history_col_idx).value = new_hist

    wb.save(inventory_path)

    return (success_count, failed_list)


# -------------------------
# Tkinter UI
# -------------------------
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Сливане на поръчка + цени (Excel)")
        self.geometry("1200x700")

        self.order_path = tk.StringVar(value="")
        self.prices_path = tk.StringVar(value="")
        self.protocols_dir_var = tk.StringVar(value="(не е избрана)")
        self.inventory_path = tk.StringVar(value="(не е избран)")

        self.df_merged = None
        self._rendered_index_map = []
        self._current_file_path = None  # Path to currently loaded file for saving
        self._multiple_prices_paths = []  # List of multiple price files

        self._build_ui()

    def _build_ui(self):
        top = ttk.Frame(self, padding=10)
        top.pack(side=tk.TOP, fill=tk.X)

        btn_order = ttk.Button(top, text="Качи Поръчка (.xls/.xlsx)", command=self.pick_order)
        btn_prices = ttk.Button(top, text="Качи Цени (.xls/.xlsx)", command=self.pick_multiple_prices)
        btn_merge = ttk.Button(top, text="Слей", command=self.do_merge)
        btn_save = ttk.Button(top, text="Запази като...", command=self.save_xlsx)

        self.search_var = tk.StringVar(value="")
        self.search_entry = ttk.Entry(top, textvariable=self.search_var, width=30)
        btn_search = ttk.Button(top, text="Търси", command=self.on_search)

        # Row 1 buttons - single order processing
        btn_order.grid(row=0, column=0, padx=5, pady=5, sticky="w")
        btn_prices.grid(row=0, column=1, padx=5, pady=5, sticky="w")
        btn_merge.grid(row=0, column=2, padx=5, pady=5, sticky="w")
        btn_save.grid(row=0, column=3, padx=5, pady=5, sticky="w")
        self.search_entry.grid(row=0, column=4, padx=5, pady=5, sticky="w")
        btn_search.grid(row=0, column=5, padx=5, pady=5, sticky="w")

        # Row 2 - protocol management buttons
        btn_choose_protocols = ttk.Button(top, text="Избери папка за протоколи", command=self.choose_protocols_folder)
        btn_batch = ttk.Button(top, text="Качи много поръчки", command=self.batch_process)
        btn_view_protocols = ttk.Button(top, text="Преглед протоколи", command=self.view_protocols)
        btn_close_protocol = ttk.Button(top, text="Приключи протокол", command=self.close_protocol)
        btn_reopen_protocol = ttk.Button(top, text="Отвори протокол", command=self.reopen_protocol)
        
        btn_choose_protocols.grid(row=1, column=0, padx=5, pady=2, sticky="w")
        btn_batch.grid(row=1, column=1, padx=5, pady=2, sticky="w")
        btn_view_protocols.grid(row=1, column=2, padx=5, pady=2, sticky="w")
        btn_close_protocol.grid(row=1, column=3, padx=5, pady=2, sticky="w")
        btn_reopen_protocol.grid(row=1, column=4, padx=5, pady=2, sticky="w")

        # Row 3 - inventory buttons
        btn_load_inventory = ttk.Button(top, text="Качи наличности", command=self.pick_inventory)
        btn_check_inventory = ttk.Button(top, text="📦 Провери и извади от склад", command=self.check_inventory_ui)
        
        btn_load_inventory.grid(row=2, column=0, padx=5, pady=2, sticky="w")
        btn_check_inventory.grid(row=2, column=1, padx=5, pady=2, sticky="w")

        ttk.Label(top, text="Поръчка:").grid(row=3, column=0, sticky="w")
        ttk.Label(top, textvariable=self.order_path).grid(row=3, column=1, columnspan=6, sticky="w")

        ttk.Label(top, text="Цени:").grid(row=4, column=0, sticky="w")
        ttk.Label(top, textvariable=self.prices_path).grid(row=4, column=1, columnspan=6, sticky="w")

        ttk.Label(top, text="Протоколи:").grid(row=5, column=0, sticky="w")
        ttk.Label(top, textvariable=self.protocols_dir_var).grid(row=5, column=1, columnspan=6, sticky="w")

        ttk.Label(top, text="Наличности:").grid(row=6, column=0, sticky="w")
        ttk.Label(top, textvariable=self.inventory_path).grid(row=6, column=1, columnspan=6, sticky="w")

        mid = ttk.Frame(self, padding=(10, 0, 10, 10))
        mid.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(mid, show="headings")
        vsb = ttk.Scrollbar(mid, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(mid, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        self.tree.bind("<Double-1>", self.on_row_double_click)

        # Bottom bar with status
        self.status = tk.StringVar(value="Избери двата файла и натисни 'Слей'.")
        ttk.Label(self, textvariable=self.status, padding=10).pack(side=tk.BOTTOM, fill=tk.X)

    def pick_order(self):
        path = filedialog.askopenfilename(
            title="Избери файл Поръчка",
            filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
        )
        if path:
            self.order_path.set(path)

    def pick_inventory(self):
        """Select inventory file with available labels."""
        path = filedialog.askopenfilename(
            title="Избери файл с наличности",
            filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
        )
        if path:
            self.inventory_path.set(path)
            self.status.set(f"Зареден файл с наличности: {Path(path).name}")

    def check_inventory_ui(self):
        """Check current order against inventory - show rich popup with colored table."""
        if self.df_merged is None or self.df_merged.empty:
            messagebox.showwarning("Няма данни", "Първо зареди и слей поръчка.")
            return

        inv_path = self.inventory_path.get()
        if inv_path == "(не е избран)" or not Path(inv_path).exists():
            messagebox.showwarning("Наличности", "Първо качи файл с наличности.")
            return

        try:
            df_inventory = load_inventory(inv_path)
            self.df_merged = check_inventory(self.df_merged, df_inventory)
            self._load_table(self.df_merged)
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при проверка: {e}")
            return

        # --- Build popup ---
        popup = tk.Toplevel(self)
        popup.title("Проверка на наличности")
        popup.geometry("1000x640")
        popup.minsize(750, 480)

        # Header
        hdr = ttk.Frame(popup, padding=(15, 12, 15, 6))
        hdr.pack(fill=tk.X)
        ttk.Label(hdr, text="📦  Проверка на наличности", font=("", 15, "bold")).pack(side=tk.LEFT)

        # Summary badges
        def _get_status_rows(val):
            if 'Статус наличност' not in self.df_merged.columns:
                return pd.DataFrame()
            s = self.df_merged['Статус наличност'].fillna('').astype(str)
            if val == 'low':
                return self.df_merged[s.str.startswith('Недостатъчно')]
            if val == 'difftl':
                return self.df_merged[s.str.startswith('Различен ТЛ')]
            return self.df_merged[s == val]

        ok_rows       = _get_status_rows('Достатъчно')
        low_rows      = _get_status_rows('low')
        none_rows     = _get_status_rows('Няма')
        difftl_rows   = _get_status_rows('difftl')
        nostock_rows  = _get_status_rows('Не в склада')
        no_tl_rows    = _get_status_rows('')

        badge_frame = ttk.Frame(hdr)
        badge_frame.pack(side=tk.RIGHT)
        tk.Label(badge_frame, text=f"✅ Достатъчно: {len(ok_rows)}",        bg="#d4edda", fg="#155724", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Label(badge_frame, text=f"⚠️  Недостатъчно: {len(low_rows)}",    bg="#fff3cd", fg="#856404", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Label(badge_frame, text=f"🔴 Няма: {len(none_rows)}",            bg="#f8d7da", fg="#721c24", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Label(badge_frame, text=f"🟣 Различен ТЛ: {len(difftl_rows)}",  bg="#e8d5f5", fg="#5a1e7a", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Label(badge_frame, text=f"🔵 Не в склада: {len(nostock_rows)}", bg="#cce5ff", fg="#004085", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)
        tk.Label(badge_frame, text=f"⬜ Без ТЛ: {len(no_tl_rows)}",         bg="#e2e3e5", fg="#383d41", padx=8, pady=3, font=("", 10, "bold")).pack(side=tk.LEFT, padx=2)

        ttk.Separator(popup, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=10)

        # Filter buttons
        filter_frame = ttk.Frame(popup, padding=(10, 6))
        filter_frame.pack(fill=tk.X)
        ttk.Label(filter_frame, text="Покажи:").pack(side=tk.LEFT, padx=(0, 6))
        filter_var = tk.StringVar(value="Всички")

        def apply_filter(val):
            filter_var.set(val)
            refresh_table()

        for label in ["Всички", "Достатъчно", "Недостатъчно", "Няма", "Различен ТЛ", "Не в склада"]:
            ttk.Button(filter_frame, text=label, width=14,
                       command=lambda l=label: apply_filter(l)).pack(side=tk.LEFT, padx=2)

        # ── Select-all checkbox in header ─────────────────────────────────
        select_all_var = tk.BooleanVar(value=False)

        def _toggle_all():
            state = select_all_var.get()
            for iid in tree.get_children():
                rd = row_data.get(iid)
                if rd is None:
                    continue
                rd["checked"] = state
                _update_check_cell(iid, state)
            _update_deduct_btn()

        # Table  ── columns: ☐ | Артикул | ТЛ | Поръчани | Налични | Остатък | Статус | Бр. за изв.
        cols = ("☐", "Артикул", "ТЛ", "Поръчани", "Налични", "Остатък", "Статус", "Бр. за изв.")
        table_frame = ttk.Frame(popup)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(4, 0))

        tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=20)
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        col_widths = {"☐": 36, "Артикул": 210, "ТЛ": 66, "Поръчани": 84,
                      "Налични": 84, "Остатък": 84, "Статус": 170, "Бр. за изв.": 90}
        for c in cols:
            tree.heading(c, text=c, anchor="center")
            tree.column(c, width=col_widths.get(c, 100),
                        anchor="w" if c == "Артикул" else "center",
                        stretch=(c == "Артикул"))

        # clicking the ☐ header = select/deselect all
        tree.heading("☐", text="☐", command=_toggle_all)

        tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)

        # Row color tags  (two variants: normal + checked)
        _TAG_COLORS = {
            "ok":      ("#d4edda", "#155724"),
            "low":     ("#fff3cd", "#856404"),
            "none":    ("#f8d7da", "#721c24"),
            "difftl":  ("#e8d5f5", "#5a1e7a"),
            "nostock": ("#cce5ff", "#004085"),
            "notl":    ("#f8f9fa", "#6c757d"),
        }
        _CHECKED_BG = "#b8daff"  # bright blue highlight when checked

        for tag, (bg, fg) in _TAG_COLORS.items():
            tree.tag_configure(tag,              background=bg, foreground=fg)
            tree.tag_configure(tag + "_checked", background=_CHECKED_BG, foreground="#003070")

        # row_data: iid → {df_idx, art, tl, ordered, avail, tag, checked, qty_var}
        row_data: dict = {}

        def _update_check_cell(iid, checked: bool):
            """Update the ☐/☑ cell and row highlight for one iid."""
            rd  = row_data.get(iid)
            if rd is None:
                return
            tag = rd["tag"] + ("_checked" if checked else "")
            vals = list(tree.item(iid, "values"))
            vals[0] = "☑" if checked else "☐"
            tree.item(iid, values=vals, tags=(tag,))

        # Holder for btn_deduct (defined later, but referenced in _update_deduct_btn)
        _widgets = {"btn_deduct": None}

        def _update_deduct_btn():
            """Enable/disable the deduct button based on selection."""
            btn = _widgets.get("btn_deduct")
            if btn is None:
                return
            n = sum(1 for rd in row_data.values() if rd["checked"])
            if n:
                btn.config(state="normal", text=f"✂️  Извади избраните ({n})")
            else:
                btn.config(state="disabled", text="✂️  Извади избраните")

        def refresh_table():
            tree.delete(*tree.get_children())
            row_data.clear()
            flt = filter_var.get()
            for idx, row in self.df_merged.iterrows():
                art     = str(row.get("Артикул", ""))
                tl      = str(row.get("Технологичен лист", "")).strip()
                ordered = row.get("Бройки", 0)
                avail   = row.get("Налични", "")
                status  = str(row.get("Статус наличност", "")).strip()

                try:
                    remainder = int(avail) - int(ordered)
                    remainder_str = str(remainder)
                except Exception:
                    remainder_str = ""

                if status == "Достатъчно":
                    tag = "ok"
                elif status.startswith("Недостатъчно"):
                    tag = "low"
                elif status == "Няма":
                    tag = "none"
                elif status.startswith("Различен ТЛ"):
                    tag = "difftl"
                elif status == "Не в склада":
                    tag = "nostock"
                else:
                    tag = "notl"

                if flt == "Достатъчно"   and tag != "ok":      continue
                if flt == "Недостатъчно" and tag != "low":     continue
                if flt == "Няма"         and tag != "none":    continue
                if flt == "Различен ТЛ"  and tag != "difftl":  continue
                if flt == "Не в склада"  and tag != "nostock": continue

                avail_str = str(int(avail)) if avail != "" and avail == avail else "-"
                display_status = status if status else "Без ТЛ"

                # default deduct qty = min(available, ordered)
                try:
                    default_qty = str(max(0, min(int(avail_str), int(ordered))))
                except Exception:
                    default_qty = str(ordered)

                iid = tree.insert("", "end",
                    values=("☐", art, tl if tl else "-", ordered,
                            avail_str, remainder_str, display_status, default_qty),
                    tags=(tag,))
                row_data[iid] = {
                    "df_idx":  idx,
                    "art":     art,
                    "tl":      tl,
                    "ordered": ordered,
                    "avail":   avail_str,
                    "tag":     tag,
                    "checked": False,
                    "qty":     default_qty,   # current value shown in col 8
                }
            _update_deduct_btn()

        refresh_table()

        # ── Inline editing of "Бр. за изв." column ───────────────────────
        _qty_entry_widget = [None]   # holds the active Entry widget

        def _hide_qty_entry(save=True):
            w = _qty_entry_widget[0]
            if w is None:
                return
            iid, var = w._iid, w._var
            if save:
                val = var.get().strip()
                rd  = row_data.get(iid)
                if rd is not None:
                    rd["qty"] = val
                    vals = list(tree.item(iid, "values"))
                    vals[7] = val
                    tree.item(iid, values=vals)
            w.destroy()
            _qty_entry_widget[0] = None

        def _show_qty_entry(iid):
            _hide_qty_entry(save=True)
            rd = row_data.get(iid)
            if rd is None:
                return
            # Get bounding box of the "Бр. за изв." cell (column index 7, i.e. #8)
            bbox = tree.bbox(iid, column="#8")
            if not bbox:
                return
            x, y, w, h = bbox
            var = tk.StringVar(value=rd["qty"])
            entry = tk.Entry(tree, textvariable=var, justify="center",
                             font=("", 10), relief="solid", bd=1)
            entry._iid = iid
            entry._var = var
            entry.place(x=x, y=y, width=w, height=h)
            entry.focus_set()
            entry.select_range(0, tk.END)
            entry.bind("<Return>",  lambda e: _hide_qty_entry(save=True))
            entry.bind("<Tab>",     lambda e: _hide_qty_entry(save=True))
            entry.bind("<Escape>",  lambda e: _hide_qty_entry(save=False))
            entry.bind("<FocusOut>", lambda e: _hide_qty_entry(save=True))
            _qty_entry_widget[0] = entry

        # ── Click handler ─────────────────────────────────────────────────
        def _on_click(event):
            iid = tree.identify_row(event.y)
            col = tree.identify_column(event.x)
            if not iid:
                _hide_qty_entry(save=True)
                return

            if col == "#1":               # ── checkbox column ──
                _hide_qty_entry(save=True)
                rd = row_data.get(iid)
                if rd is None:
                    return
                rd["checked"] = not rd["checked"]
                _update_check_cell(iid, rd["checked"])
                _update_deduct_btn()
                return

            if col == "#8":               # ── Бр. за изв. column ──
                _show_qty_entry(iid)
                return

            _hide_qty_entry(save=True)

        tree.bind("<ButtonRelease-1>", _on_click)

        # ── Bottom bar ────────────────────────────────────────────────────
        ttk.Separator(popup, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=10, pady=(6, 0))
        bottom = ttk.Frame(popup, padding=(10, 8))
        bottom.pack(fill=tk.X)

        def _do_deduct_selected():
            _hide_qty_entry(save=True)
            inv_path_val = self.inventory_path.get()
            if inv_path_val == "(не е избран)" or not Path(inv_path_val).exists():
                messagebox.showwarning("Наличности", "Първо качи файл с наличности.", parent=popup)
                return

            selected = {iid: rd for iid, rd in row_data.items() if rd["checked"]}
            if not selected:
                return

            # Validate quantities first
            errors = []
            for iid, rd in selected.items():
                try:
                    q = int(rd["qty"])
                    if q <= 0:
                        raise ValueError
                except ValueError:
                    errors.append(f"{rd['art']}: невалидно количество '{rd['qty']}'")
            if errors:
                messagebox.showerror("Грешка", "\n".join(errors), parent=popup)
                return

            # Confirm
            lines = "\n".join(
                f"  • {rd['art']}  →  {rd['qty']} бр."
                for rd in selected.values()
            )
            if not messagebox.askyesno(
                    "Потвърждение",
                    f"Ще се приспаднат от склада:\n\n{lines}\n\nПродължаваш?",
                    parent=popup):
                return

            # Build combined single-row-per-item DataFrame and call reserve_inventory
            rows_to_reserve = []
            for iid, rd in selected.items():
                df_idx  = rd["df_idx"]
                one_row = self.df_merged.loc[[df_idx]].copy()
                one_row["Бројки"] = int(rd["qty"])   # override with chosen qty
                one_row["Бройки"] = int(rd["qty"])
                rows_to_reserve.append(one_row)

            combined_df = pd.concat(rows_to_reserve, ignore_index=True)

            try:
                success, failed = reserve_inventory(combined_df, inv_path_val)
            except Exception as ex:
                messagebox.showerror("Грешка", str(ex), parent=popup)
                return

            # Refresh inventory data
            try:
                df_inv_fresh = load_inventory(inv_path_val)
                self.df_merged = check_inventory(self.df_merged, df_inv_fresh)
                self._load_table(self.df_merged)
            except Exception:
                pass
            refresh_table()

            msg = f"✅ Приспаднати {success} артикула от склада."
            if failed:
                msg += f"\n\nПроблеми:\n" + "\n".join(f"  • {f}" for f in failed[:10])
            messagebox.showinfo("Извадено от склада", msg, parent=popup)

        btn_deduct = ttk.Button(bottom, text="✂️  Извади избраните",
                                state="disabled", command=_do_deduct_selected)
        btn_deduct.pack(side=tk.LEFT, padx=(0, 10))
        _widgets["btn_deduct"] = btn_deduct

        ttk.Label(bottom, text="☐ = избери ред  |  кликни 'Бр. за изв.' за да промениш количеството",
                  foreground="#555").pack(side=tk.LEFT, padx=4)
        ttk.Button(bottom, text="Затвори", command=popup.destroy).pack(side=tk.RIGHT, padx=5)

        self.status.set("Наличностите са проверени")

    def reserve_inventory_ui(self):
        """Reserve items from inventory based on current order."""
        if self.df_merged is None or self.df_merged.empty:
            messagebox.showwarning("Няма данни", "Първо зареди и слей поръчка.")
            return

        inv_path = self.inventory_path.get()
        if inv_path == "(не е избран)" or not Path(inv_path).exists():
            messagebox.showwarning("Наличности", "Първо качи файл с наличности.")
            return

        # Confirm action
        if not messagebox.askyesno("Потвърждение",
            "Това ще намали наличностите в склада.\n\n"
            "Сигурен ли си, че искаш да извадиш?"):
            return

        try:
            success, failed = reserve_inventory(self.df_merged, inv_path)

            msg = f"Извадени: {success} артикула\n"
            if failed:
                msg += f"\nПроблеми ({len(failed)}):\n"
                for f in failed[:10]:
                    msg += f"  • {f}\n"
                if len(failed) > 10:
                    msg += f"  ... и още {len(failed) - 10}"

            messagebox.showinfo("Извадено от склад", msg)

            # Refresh inventory check
            self.check_inventory_ui()

            self.status.set(f"Извадени {success} артикула от склада")
        except Exception as e:
            messagebox.showerror("Грешка", f"Грешка при резервиране: {e}")

    def pick_multiple_prices(self):
        """Open dialog to select multiple price files."""
        popup = tk.Toplevel(self)
        popup.title("Качи цени")
        popup.geometry("600x520")
        popup.transient(self)
        popup.minsize(500, 420)
        
        price_files = list(self._multiple_prices_paths)  # Copy existing
        
        # Title
        ttk.Label(popup, text="Качи файлове с цени", font=("", 14, "bold")).pack(pady=(15, 5))
        ttk.Label(popup, text="Цените ще бъдат обединени в един ценоразпис").pack(pady=(0, 5))
        ttk.Label(popup, text="⚠️ По-късно добавените файлове имат приоритет при дублиращи се артикули", 
                  foreground="orange").pack(pady=(0, 10))
        
        # Add files button
        def add_files_dialog():
            paths = filedialog.askopenfilenames(
                title="Избери файлове с цени",
                filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
            )
            if paths:
                for p in paths:
                    p_str = str(p).strip()
                    if p_str.lower().endswith(('.xls', '.xlsx')) and p_str not in price_files:
                        price_files.append(p_str)
                        files_listbox.insert(tk.END, Path(p_str).name)
                update_label()
        
        add_btn = ttk.Button(popup, text="📂 Добави файлове...", command=add_files_dialog)
        add_btn.pack(pady=10)
        
        # Files list label
        ttk.Label(popup, text="Добавени ценови файлове:").pack(anchor="w", padx=20)
        
        # Listbox for files
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        files_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, height=8)
        files_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=files_listbox.yview)
        files_listbox.configure(yscrollcommand=files_scrollbar.set)
        
        files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        files_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Load existing files into listbox
        for p in price_files:
            files_listbox.insert(tk.END, Path(p).name)
        
        # Status label
        status_label = ttk.Label(popup, text="Няма добавени файлове", foreground="gray")
        status_label.pack(pady=5)
        
        def update_label():
            if price_files:
                status_label.configure(text=f"{len(price_files)} файла добавени", foreground="green")
            else:
                status_label.configure(text="Няма добавени файлове", foreground="gray")
        
        update_label()
        
        # Action buttons for list management
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=20, pady=5)
        
        def remove_selected():
            selected = list(files_listbox.curselection())
            for i in reversed(selected):
                files_listbox.delete(i)
                del price_files[i]
            update_label()
        
        def clear_all():
            files_listbox.delete(0, tk.END)
            price_files.clear()
            update_label()
        
        ttk.Button(btn_frame, text="Премахни избраните", command=remove_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Изчисти всички", command=clear_all).pack(side=tk.LEFT, padx=5)
        
        # Separator and bottom buttons - FIXED at bottom
        ttk.Separator(popup, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=20, pady=10, side=tk.BOTTOM)
        
        bottom_frame = ttk.Frame(popup)
        bottom_frame.pack(fill=tk.X, padx=20, pady=10, side=tk.BOTTOM)
        
        def do_confirm():
            if not price_files:
                messagebox.showwarning("Липсват файлове", "Добави поне един файл с цени.")
                return
            self._multiple_prices_paths = list(price_files)
            self.prices_path.set(f"[{len(price_files)} файла]")
            popup.destroy()
            self.status.set(f"Заредени {len(price_files)} файла с цени")
        
        ttk.Button(bottom_frame, text="Отказ", command=popup.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(bottom_frame, text="Потвърди", command=do_confirm).pack(side=tk.RIGHT, padx=5)

    def choose_protocols_folder(self):
        path = filedialog.askdirectory(title="Избери папка за протоколи")
        if not path:
            return
        try:
            set_protocols_dir(path)
            self.protocols_dir_var.set(str(PROTOCOLS_DIR))
            self.status.set(f"Папка за протоколи: {PROTOCOLS_DIR}")
        except Exception as e:
            messagebox.showerror("Грешка", f"Не мога да задам папката: {e}")

    def view_protocols(self):
        """Show list of all protocols with their status."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        protocols = []
        for p in PROTOCOLS_DIR.glob("protocol_*.xlsx"):
            name = p.stem
            # Check if closed by name (_CLOSED suffix)
            if "_CLOSED" in name:
                status = "ПРИКЛЮЧЕН"
                display_name = name.replace("_CLOSED", "")
            else:
                status = "Отворен"
                display_name = name
            try:
                df = pd.read_excel(p, engine="openpyxl")
                rows = len(df)
            except Exception:
                rows = "?"
            protocols.append((display_name, status, rows, name))
        
        if not protocols:
            messagebox.showinfo("Протоколи", "Няма намерени протоколи в избраната папка.")
            return
        
        # Create popup window
        popup = tk.Toplevel(self)
        popup.title("Протоколи")
        popup.geometry("600x400")
        
        tree = ttk.Treeview(popup, columns=("Име", "Статус", "Редове"), show="headings")
        tree.heading("Име", text="Протокол")
        tree.heading("Статус", text="Статус")
        tree.heading("Редове", text="Редове")
        tree.column("Име", width=300)
        tree.column("Статус", width=120)
        tree.column("Редове", width=80)
        
        for display_name, status, rows, full_name in sorted(protocols, reverse=True):
            tree.insert("", "end", values=(display_name, status, rows), tags=(full_name,))
        
        tree.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        def open_selected():
            sel = tree.selection()
            if not sel:
                return
            # Get full name from tags
            full_name = tree.item(sel[0])["tags"][0]
            prot_path = PROTOCOLS_DIR / f"{full_name}.xlsx"
            if prot_path.exists():
                try:
                    df = pd.read_excel(prot_path, engine="openpyxl")
                    self.df_merged = df
                    self._current_file_path = str(prot_path)  # Remember path for saving
                    self._load_table(df)
                    self.status.set(f"Заредени {len(df)} реда от {full_name} (двоен клик за редакция)")
                    popup.destroy()
                except Exception as e:
                    messagebox.showerror("Грешка", f"Не мога да отворя протокола: {e}")
        
        btn_open = ttk.Button(popup, text="Отвори в таблицата", command=open_selected)
        btn_open.pack(pady=5)

    def close_protocol(self):
        """Mark a protocol as closed (no more rows can be added)."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        # Find protocols without _CLOSED in name
        open_protocols = [p.stem for p in PROTOCOLS_DIR.glob("protocol_*.xlsx") 
                          if "_CLOSED" not in p.stem]
        
        if not open_protocols:
            messagebox.showinfo("Протоколи", "Няма отворени протоколи за приключване.")
            return
        
        # Create selection dialog
        popup = tk.Toplevel(self)
        popup.title("Приключи протокол")
        popup.geometry("400x300")
        
        ttk.Label(popup, text="Избери протокол за приключване:").pack(pady=10)
        
        listbox = tk.Listbox(popup, selectmode=tk.SINGLE, width=50, height=10)
        for p in sorted(open_protocols, reverse=True):
            listbox.insert(tk.END, p)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        def do_close():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Избор", "Избери протокол от списъка.")
                return
            name = listbox.get(sel[0])
            
            if messagebox.askyesno("Потвърждение", 
                f"Сигурен ли си, че искаш да приключиш протокол '{name}'?\n\n"
                "След приключване няма да можеш да добавяш нови редове към него.\n"
                "Файлът ще бъде преименуван с _CLOSED и защитен от промени."):
                
                prot_file = PROTOCOLS_DIR / f"{name}.xlsx"
                closed_file = PROTOCOLS_DIR / f"{name}_CLOSED.xlsx"
                
                if prot_file.exists():
                    # Rename to _CLOSED
                    try:
                        prot_file.rename(closed_file)
                        # Set file as read-only
                        set_file_readonly(closed_file, readonly=True)
                        messagebox.showinfo("Готово", f"Протокол '{name}' е приключен.\nНово име: {closed_file.name}")
                    except Exception as e:
                        messagebox.showerror("Грешка", f"Не мога да преименувам файла: {e}")
                        return
                
                popup.destroy()
        
        ttk.Button(popup, text="Приключи", command=do_close).pack(pady=10)

    def reopen_protocol(self):
        """Reopen a closed protocol."""
        if self.protocols_dir_var.get() == "(не е избрана)":
            messagebox.showwarning("Протоколи", "Първо избери папка за протоколи.")
            return
        
        # Find protocols with _CLOSED in name
        closed_protocols = [p.stem for p in PROTOCOLS_DIR.glob("protocol_*_CLOSED.xlsx")]
        
        if not closed_protocols:
            messagebox.showinfo("Протоколи", "Няма приключени протоколи за отваряне.")
            return
        
        # Create selection dialog
        popup = tk.Toplevel(self)
        popup.title("Отвори протокол")
        popup.geometry("400x300")
        
        ttk.Label(popup, text="Избери протокол за отваряне:").pack(pady=10)
        
        listbox = tk.Listbox(popup, selectmode=tk.SINGLE, width=50, height=10)
        for p in sorted(closed_protocols, reverse=True):
            listbox.insert(tk.END, p)
        listbox.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        
        def do_reopen():
            sel = listbox.curselection()
            if not sel:
                messagebox.showwarning("Избор", "Избери протокол от списъка.")
                return
            name = listbox.get(sel[0])  # e.g. protocol_2026_w7_CLOSED
            
            closed_file = PROTOCOLS_DIR / f"{name}.xlsx"
            # Remove _CLOSED from name
            open_name = name.replace("_CLOSED", "")
            open_file = PROTOCOLS_DIR / f"{open_name}.xlsx"
            
            if closed_file.exists():
                try:
                    # Remove read-only flag first
                    set_file_readonly(closed_file, readonly=False)
                    # Rename back
                    closed_file.rename(open_file)
                    messagebox.showinfo("Готово", f"Протокол '{open_name}' е отворен за добавяне на редове.")
                except Exception as e:
                    messagebox.showerror("Грешка", f"Не мога да преименувам файла: {e}")
                    return
            
            popup.destroy()
        
        ttk.Button(popup, text="Отвори", command=do_reopen).pack(pady=10)

    def batch_process(self):
        """Process multiple order files at once and add them to weekly protocols."""
        if self.protocols_dir_var.get() == "(не е избрана)" or PROTOCOLS_DIR is None:
            messagebox.showwarning("Папка за протоколи", "Първо избери папка за протоколи.")
            return
        
        # Create window
        popup = tk.Toplevel(self)
        popup.title("Качи много поръчки")
        popup.geometry("600x520")
        popup.transient(self)
        popup.minsize(500, 420)
        
        order_files = []
        
        # Title
        ttk.Label(popup, text="Качи поръчки към протоколи", font=("", 14, "bold")).pack(pady=(15, 10))
        
        # Add files button - PROMINENT at top
        def add_files_dialog():
            paths = filedialog.askopenfilenames(
                title="Избери поръчки",
                filetypes=[("Excel", "*.xlsx"), ("Excel 97-2003", "*.xls"), ("All files", "*.*")]
            )
            if paths:
                for p in paths:
                    p_str = str(p).strip()
                    if p_str.lower().endswith(('.xls', '.xlsx')) and p_str not in order_files:
                        order_files.append(p_str)
                        files_listbox.insert(tk.END, Path(p_str).name)
                update_label()
        
        add_btn = ttk.Button(popup, text="📂 Добави файлове...", command=add_files_dialog)
        add_btn.pack(pady=10)
        
        # Files list label
        ttk.Label(popup, text="Добавени поръчки:").pack(anchor="w", padx=20)
        
        # Listbox for files
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)
        
        files_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, height=12)
        files_scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=files_listbox.yview)
        files_listbox.configure(yscrollcommand=files_scrollbar.set)
        
        files_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        files_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Status label
        status_label = ttk.Label(popup, text="Няма добавени файлове", foreground="gray")
        status_label.pack(pady=5)
        
        def update_label():
            if order_files:
                status_label.configure(text=f"{len(order_files)} файла добавени", foreground="green")
            else:
                status_label.configure(text="Няма добавени файлове", foreground="gray")
        
        # Action buttons
        btn_frame = ttk.Frame(popup)
        btn_frame.pack(fill=tk.X, padx=20, pady=5)
        
        def remove_selected():
            selected = list(files_listbox.curselection())
            for i in reversed(selected):
                files_listbox.delete(i)
                del order_files[i]
            update_label()
        
        def clear_all():
            files_listbox.delete(0, tk.END)
            order_files.clear()
            update_label()
        
        ttk.Button(btn_frame, text="Премахни избраните", command=remove_selected).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Изчисти всички", command=clear_all).pack(side=tk.LEFT, padx=5)
        
        # Separator and bottom buttons
        ttk.Separator(popup, orient=tk.HORIZONTAL).pack(fill=tk.X, padx=20, pady=10)
        
        bottom_frame = ttk.Frame(popup)
        bottom_frame.pack(fill=tk.X, padx=20, pady=10)
        
        def do_process():
            if not order_files:
                messagebox.showwarning("Липсват файлове", "Добави поне една поръчка.")
                return
            popup.destroy()
            self._process_batch_files(order_files)
        
        ttk.Button(bottom_frame, text="Отказ", command=popup.destroy).pack(side=tk.RIGHT, padx=5)
        ttk.Button(bottom_frame, text="Обработи", command=do_process).pack(side=tk.RIGHT, padx=5)

    def _process_batch_files(self, order_files):
        """Process the batch of order files (already contain prices)."""
        processed = 0
        errors = []
        all_merged = []
        
        # Expected columns in protocol
        protocol_cols = ["Артикул", "Размер", "Бройки", "Ед. Цена", "Сума", 
                         "Номер на поръчка и ред", "Дата на доставка", "Технологичен лист", "Материал"]
        
        # Column name mappings (various formats -> standard format)
        col_mappings = {
            # Артикул
            "артикул": "Артикул",
            "item": "Артикул",
            "item number": "Артикул",
            "име на артикул": "Артикул",
            "продукт": "Артикул",
            # Размер
            "размер": "Размер",
            "size": "Размер",
            # Бройки
            "бройки": "Бройки",
            "брой": "Бройки",
            "qty": "Бройки",
            "quantity": "Бройки",
            "количество": "Бройки",
            # Ед. Цена
            "ед. цена": "Ед. Цена",
            "ед цена": "Ед. Цена",
            "единична цена": "Ед. Цена",
            "unit price": "Ед. Цена",
            "цена": "Ед. Цена",
            # Сума
            "сума": "Сума",
            "total": "Сума",
            "amount": "Сума",
            "обща сума": "Сума",
            # Номер на поръчка и ред
            "номер на поръчка и ред": "Номер на поръчка и ред",
            "номер поръчка": "Номер на поръчка и ред",
            "поръчка": "Номер на поръчка и ред",
            "order": "Номер на поръчка и ред",
            "order number": "Номер на поръчка и ред",
            "purchase order": "Номер на поръчка и ред",
            # Дата на доставка
            "дата на доставка": "Дата на доставка",
            "дата доставка": "Дата на доставка",
            "дата": "Дата на доставка",
            "delivery date": "Дата на доставка",
            "date": "Дата на доставка",
            # Технологичен лист
            "технологичен лист": "Технологичен лист",
            "тл": "Технологичен лист",
            "tech sheet": "Технологичен лист",
            # Материал
            "материал": "Материал",
            "material": "Материал",
        }
        
        # Required columns - file must have at least these
        required_cols = ["Артикул", "Бройки"]
        
        for order_path in order_files:
            try:
                # Read the order file directly (it already has all data including prices)
                df = read_excel_any(order_path)
                
                if df.empty:
                    errors.append(f"{Path(order_path).name}: Файлът е празен")
                    continue
                
                # Normalize column names
                new_columns = {}
                for col in df.columns:
                    col_lower = str(col).strip().lower()
                    if col_lower in col_mappings:
                        new_columns[col] = col_mappings[col_lower]
                    elif str(col).strip() in protocol_cols:
                        new_columns[col] = str(col).strip()
                
                if new_columns:
                    df = df.rename(columns=new_columns)
                
                # Check if file has required columns
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    errors.append(f"{Path(order_path).name}: Липсват колони: {', '.join(missing_cols)}")
                    continue
                
                # Get order name from filename
                order_no = Path(order_path).stem
                
                # Ensure all protocol columns exist (add missing optional ones as empty)
                for col in protocol_cols:
                    if col not in df.columns:
                        df[col] = ""
                
                # Group by week and append to protocols
                ensure_dirs()
                groups = {}
                for _, row in df.iterrows():
                    wk = week_key_from_date(row.get("Дата на доставка"))
                    groups.setdefault(wk, []).append(row.to_dict())
                
                for wk, rows in groups.items():
                    df_rows = pd.DataFrame(rows)
                    try:
                        append_to_protocol(wk, df_rows, Path(order_path).name)
                    except RuntimeError as e:
                        # Protocol is closed
                        errors.append(f"{order_no}: {e}")
                
                processed += 1
                all_merged.append(df)
                
            except Exception as e:
                errors.append(f"{Path(order_path).name}: {e}")
        
        # Show results
        if all_merged:
            combined = pd.concat(all_merged, ignore_index=True)
            self.df_merged = combined
            self._load_table(combined)
        
        msg = f"Обработени {processed} от {len(order_files)} поръчки."
        if errors:
            msg += f"\n\nГрешки:\n" + "\n".join(errors[:10])
            if len(errors) > 10:
                msg += f"\n... и още {len(errors) - 10} грешки"
            messagebox.showwarning("Резултат", msg)
        else:
            messagebox.showinfo("Готово", msg)
        
        self.status.set(msg.split("\n")[0])

    def do_merge(self):
        op = self.order_path.get().strip()
        pp = self.prices_path.get().strip()
        if not op or not pp:
            messagebox.showwarning("Липсват файлове", "Моля избери и двата файла (Поръчка и Цени).")
            return

        # Изискване: първо трябва да са качени наличности
        inv_path = self.inventory_path.get()
        if inv_path == "(не е избран)" or not Path(inv_path).exists():
            messagebox.showwarning("Наличности", 
                "Първо качи файл с наличности,\nза да се покаже статусът на артикулите.")
            return

        try:
            # Check if we have multiple price files
            if self._multiple_prices_paths:
                # Merge multiple price files into one DataFrame
                combined_prices = self._combine_price_files(self._multiple_prices_paths)
                if combined_prices is None or combined_prices.empty:
                    messagebox.showerror("Грешка", "Не успях да обединя ценовите файлове.")
                    return
                self.df_merged = merge_order_and_prices(op, combined_prices)
            else:
                self.df_merged = merge_order_and_prices(op, pp)

            # Автоматично добавяне на колони Налични / Статус наличност
            try:
                df_inventory = load_inventory(inv_path)
                self.df_merged = check_inventory(self.df_merged, df_inventory)
            except Exception as inv_err:
                # Ако не успее, продължаваме без инвентарни колони
                print(f"Предупреждение: {inv_err}")

            self._current_file_path = None  # New merge, no file yet
            self._load_table(self.df_merged)
            self.status.set(f"Готово: {len(self.df_merged)} реда слети с наличности.")
        except Exception as e:
            messagebox.showerror("Грешка", str(e))
            self.status.set("Грешка при сливане.")

    def _combine_price_files(self, price_paths):
        """Combine multiple price files into a single DataFrame.
        
        Normalizes article column names to a common name before combining.
        """
        # Possible names for the article column
        article_col_names = ["код АЛ филтър", "артикул", "item", "item number", "код", "product"]
        target_article_col = "код АЛ филтър"  # Unified name
        
        dfs = []
        for path in price_paths:
            try:
                df = read_excel_any(path)
                
                # Find and rename article column to unified name
                for col in df.columns:
                    if str(col).strip().lower() in [n.lower() for n in article_col_names]:
                        if col != target_article_col:
                            df = df.rename(columns={col: target_article_col})
                        break
                
                dfs.append(df)
            except Exception as e:
                print(f"Грешка при четене на {path}: {e}")
                continue
        
        if not dfs:
            return None
        
        # Concatenate all dataframes
        combined = pd.concat(dfs, ignore_index=True)
        return combined

    def save_xlsx(self):
        if self.df_merged is None or self.df_merged.empty:
            messagebox.showinfo("Няма данни", "Първо натисни 'Слей'.")
            return

        default_name = "Porachka.xlsx"
        try:
            first_ref = str(self.df_merged.iloc[0]["Номер на поръчка и ред"])
            order_no = first_ref.split("-")[0]
            default_name = f"Porachka_{order_no}.xlsx"
        except Exception:
            pass

        out_path = filedialog.asksaveasfilename(
            title="Запази като",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel Workbook", "*.xlsx")]
        )
        if not out_path:
            return

        try:
            # Преобразуваме датата в реална дата за Numbers/Excel съвместимост
            df_to_save = self.df_merged.copy()
            if "Дата на доставка" in df_to_save.columns:
                df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
            
            # Ако файлът съществува, презаписваме дублиращите се редове
            if Path(out_path).exists():
                try:
                    existing = pd.read_excel(out_path, engine="openpyxl")
                    if "Номер на поръчка и ред" in existing.columns and "Номер на поръчка и ред" in df_to_save.columns:
                        # Get the order refs from new data
                        new_refs = set(df_to_save["Номер на поръчка и ред"].dropna().astype(str).tolist())
                        # Keep only rows from existing that are NOT in new data
                        existing_filtered = existing[~existing["Номер на поръчка и ред"].astype(str).isin(new_refs)]
                        df_to_save = pd.concat([existing_filtered, df_to_save], ignore_index=True)
                        # Convert date again after merge
                        if "Дата на доставка" in df_to_save.columns:
                            df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
                except Exception:
                    pass  # If can't read existing, just overwrite
            
            # Наредба на колоните: стандартните отпред, Налични/Статус накрая
            standard_cols = [c for c in df_to_save.columns
                             if c not in ("Налични", "Статус наличност")]
            inv_cols = [c for c in ("Налични", "Статус наличност")
                        if c in df_to_save.columns]
            df_to_save = df_to_save[standard_cols + inv_cols]

            with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                df_to_save.to_excel(writer, index=False, sheet_name="Porachka")

            # Оцветяване на статусните редове
            try:
                _apply_inventory_colors_xlsx(out_path)
            except Exception:
                pass
            
            # Прилагаме формат за дата
            try:
                _apply_date_format_xlsx(out_path, header_name="Дата на доставка")
            except Exception:
                pass
            
            messagebox.showinfo("Записано", f"Файлът е записан:\n{out_path}")
            self.status.set(f"Записано: {out_path}")
        except Exception as e:
            messagebox.showerror("Грешка при запис", str(e))
            self.status.set("Грешка при запис.")
            return

        # Check if protocols folder is selected - append to weekly protocols
        if self.protocols_dir_var.get() == "(не е избрана)":
            # No protocols folder selected, skip adding to protocols
            return

        # Append to weekly protocols
        try:
            ensure_dirs()
            source_name = Path(out_path).name

            # append each row to its weekly protocol
            groups = {}
            for _, row in self.df_merged.iterrows():
                wk = week_key_from_date(row.get("Дата на доставка"))
                groups.setdefault(wk, []).append(row.to_dict())

            for wk, rows in groups.items():
                df_rows = pd.DataFrame(rows)
                try:
                    append_to_protocol(wk, df_rows, source_name)
                except Exception as e:
                    messagebox.showwarning("Протокол", str(e))

        except Exception as e:
            messagebox.showwarning("Добавяне към протокол", f"Грешка: {e}")

    def _load_table(self, df: pd.DataFrame):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())

        cols = list(df.columns)
        self.tree["columns"] = cols

        for c in cols:
            self.tree.heading(c, text=c)
            self.tree.column(c, width=140, anchor="w")

        max_rows = 2000
        self._rendered_index_map = []
        idx_counter = 0

        for i, row in df.head(max_rows).iterrows():
            values = []
            for c in cols:
                v = row.get(c, "")
                try:
                    if pd.isna(v):
                        values.append("")
                        continue
                except Exception:
                    pass
                values.append(str(v))
            iid = str(idx_counter)
            self.tree.insert("", "end", iid=iid, values=values)
            self._rendered_index_map.append(i)
            idx_counter += 1

        if len(df) > max_rows:
            self.status.set(f"Показвам първите {max_rows} реда от {len(df)} (всички се записват при export).")

    def on_search(self):
        q = (self.search_var.get() or "").strip()
        if not q:
            messagebox.showinfo("Търсене", "Въведи текст за търсене (име, ТЛ или размер).")
            return
        ql = q.lower()

        def match_row(r):
            for c in ["Артикул", "Технологичен лист", "Размер"]:
                try:
                    v = str(r.get(c, "") or "").lower()
                except Exception:
                    v = ""
                if ql in v:
                    return True
            return False

        if getattr(self, 'df_merged', None) is not None and not self.df_merged.empty:
            try:
                filtered = self.df_merged[self.df_merged.apply(match_row, axis=1)]
                if not filtered.empty:
                    self.df_merged = filtered
                    self._load_table(self.df_merged)
                    self.status.set(f"Намерени {len(filtered)} реда за '{q}' (в текущото сливане).")
                    return
            except Exception:
                pass

        # No local orders directory - just search in current merge
        messagebox.showinfo("Търсене", "Няма текущо сливане за търсене. Първо заредете поръчка.")

    def on_row_double_click(self, event):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        try:
            idx = int(iid)
        except Exception:
            return
        if idx >= len(self._rendered_index_map):
            return
        df_index = self._rendered_index_map[idx]

        columns = list(self.tree['columns'])
        cur_values = [self.df_merged.at[df_index, c] if c in self.df_merged.columns else '' for c in columns]

        edit = tk.Toplevel(self)
        edit.title("Редакция на ред")

        entries = {}
        for i, c in enumerate(columns):
            ttk.Label(edit, text=c).grid(row=i, column=0, sticky='w', padx=4, pady=2)
            # Handle NaN/None values - show empty string instead of "nan"
            val = cur_values[i]
            if val is None or (isinstance(val, float) and pd.isna(val)):
                display_val = ""
            else:
                try:
                    if pd.isna(val):
                        display_val = ""
                    else:
                        display_val = str(val)
                except Exception:
                    display_val = str(val) if val is not None else ""
            v = tk.StringVar(value=display_val)
            e = ttk.Entry(edit, textvariable=v, width=60)
            e.grid(row=i, column=1, sticky='w', padx=4, pady=2)
            entries[c] = v

        def save_edit():
            for c, var in entries.items():
                val = var.get().strip()
                
                # Handle empty values
                if val == "" or val.lower() == "nan":
                    if c in ("Бройки",):
                        self.df_merged.at[df_index, c] = None
                    elif c in ("Ед. Цена", "Сума", "Технологичен лист"):
                        self.df_merged.at[df_index, c] = None
                    else:
                        self.df_merged.at[df_index, c] = ""
                elif c in ("Бройки",):
                    try:
                        vv = int(val)
                        self.df_merged.at[df_index, c] = vv
                    except Exception:
                        self.df_merged.at[df_index, c] = None
                elif c in ("Ед. Цена", "Сума", "Технологичен лист"):
                    try:
                        vv = float(str(val).replace(',', '.'))
                        self.df_merged.at[df_index, c] = vv
                    except Exception:
                        self.df_merged.at[df_index, c] = None
                else:
                    self.df_merged.at[df_index, c] = val

            edit.destroy()
            self._load_table(self.df_merged)
            
            # Auto-save to file if we have a current file path
            if self._current_file_path:
                try:
                    out_path = self._current_file_path
                    
                    # Check if file is read-only (closed protocol)
                    if is_file_readonly(Path(out_path)):
                        self.status.set("⚠️ Файлът е защитен - промените не са запазени.")
                        return
                    
                    # Save to file
                    df_to_save = self.df_merged.copy()
                    if "Дата на доставка" in df_to_save.columns:
                        df_to_save["Дата на доставка"] = pd.to_datetime(df_to_save["Дата на доставка"], errors="coerce")
                    
                    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                        df_to_save.to_excel(writer, index=False, sheet_name="Sheet1")
                    
                    try:
                        _apply_date_format_xlsx(out_path, header_name="Дата на доставка")
                    except Exception:
                        pass
                    
                    self.status.set(f"✅ Запазено в {Path(out_path).name}")
                except Exception as e:
                    self.status.set(f"⚠️ Грешка при запис: {e}")
            else:
                self.status.set("Редът е променен (използвай 'Запази като...' за да запишеш)")

        btn_save = ttk.Button(edit, text="Запиши", command=save_edit)
        btn_save.grid(row=len(columns), column=0, columnspan=2, pady=6)


if __name__ == "__main__":
    # IMPORTANT: Windows-only DPI tweak; do NOT run on macOS/Linux
    if os.name == "nt":
        try:
            import ctypes
            ctypes.windll.shcore.SetProcessDpiAwareness(1)
        except Exception:
            pass

    app = App()
    app.mainloop()
